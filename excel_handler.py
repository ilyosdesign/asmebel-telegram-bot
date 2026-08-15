# excel_handler.py

import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

ORDERS_FILE = 'buyurtmalar.xlsx'
HEADERS = [
    'Vaqt', 'Mijoz ID', 'Mijoz Username', 'Mijoz Ismi', 'Telefon Raqam', 'Mahsulot Nomi'
]

def setup_excel_file():
    if not os.path.exists(ORDERS_FILE):
        try:
            workbook = Workbook()
            sheet = workbook.active
            if sheet:
                sheet.title = "Buyurtmalar"
                sheet.append(HEADERS)
                workbook.save(ORDERS_FILE)
                logger.info(f"Excel fayli yaratildi: {ORDERS_FILE}")
        except Exception as e:
            logger.error(f"Excel fayli yaratishda xatolik: {e}")

def save_order_to_excel(user_data: dict, product_name: str = None):
    if not user_data:
        logger.warning("user_data bo'sh")
        return

    try:
        setup_excel_file()

        if not os.path.exists(ORDERS_FILE):
            logger.error(f"Excel fayli topilmadi: {ORDERS_FILE}")
            return

        workbook = load_workbook(ORDERS_FILE)
        sheet = workbook.active

        if not sheet:
            logger.error("Excel sheetini olib bo'lmadi")
            return

        order_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        product = product_name or user_data.get('Product', user_data.get('product_name', '-'))
        new_row = [
            order_time,
            str(user_data.get('User ID', user_data.get('id', '-')))[:20],
            str(user_data.get('username', '-'))[:50],
            str(user_data.get('Name', user_data.get('full_name', '-')))[:100],
            str(user_data.get('phone_number', '-'))[:20],
            str(product)[:200]
        ]
        sheet.append(new_row)
        workbook.save(ORDERS_FILE)
        logger.info(f"Buyurtma Excel ga saqlandi: {user_data.get('User ID', user_data.get('id'))}")

    except Exception as e:
        logger.error(f"Excel ga yozishda xatolik: {type(e).__name__}: {e}")